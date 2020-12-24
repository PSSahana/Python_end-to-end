import openpyxl


class HomePageData:
    get_homePageData = [{"name": "Sahana", "email": "sahana@gmail.com", "gender": "Female"},
                            {"name": "Sandesh", "email": "san@gmail.com", "gender": "Male"}]

    @staticmethod
    def get_test_data(testcaseid):
        wk = openpyxl.load_workbook("C:/Users/Sahana PS/PycharmProjects/EndtoEndFramework/TestData/data.xlsx")
        ws = wk.active
        Dict = {}
        for i in range(1, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == testcaseid:
                for j in range(2, ws.max_column + 1):
                    Dict[ws.cell(row=1, column=j).value] = ws.cell(row=i, column=j).value

        return [Dict]
