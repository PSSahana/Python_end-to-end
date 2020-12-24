import openpyxl


wk = openpyxl.load_workbook("C:/Users/Sahana PS/PycharmProjects/EndtoEndFramework/TestData/data.xlsx")
ws = wk.active
Dict = {}
for i in range(1, ws.max_row + 1):
    if ws.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, ws.max_column + 1):
            Dict[ws.cell(row=1, column=j).value] = ws.cell(row=i, column=j).value

print(Dict)